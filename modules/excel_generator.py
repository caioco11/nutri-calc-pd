"""
excel_generator.py — Geração do Relatório Excel (openpyxl)
Produz duas abas:
    1. "Tabela Nutricional — Rótulo": layout fiel ao modelo ANVISA (RDC 429/2020)
    2. "Composição Técnica — P&D": detalhamento técnico completo para uso interno
"""

import io
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, numbers
)
from openpyxl.utils import get_column_letter

# ─── Paleta de cores ───────────────────────────────────────────────────────────
COR_HEADER_ANVISA = "1F4E79"   # Azul escuro — cabeçalho "INFORMAÇÃO NUTRICIONAL"
COR_HEADER_PD     = "2E4057"   # Azul P&D
COR_ZEBRA_1       = "FFFFFF"   # Branco
COR_ZEBRA_2       = "EFF3F8"   # Azul bem claro
COR_TOTAL         = "D6E4F0"   # Azul médio para linha de totais
COR_FONTE_HEADER  = "FFFFFF"   # Branco
COR_BORDA         = "B0BEC5"   # Cinza

# ─── Estilos ──────────────────────────────────────────────────────────────────
BORDA_FINA = Border(
    left=Side(style="thin", color=COR_BORDA),
    right=Side(style="thin", color=COR_BORDA),
    top=Side(style="thin", color=COR_BORDA),
    bottom=Side(style="thin", color=COR_BORDA),
)

def _fonte(bold=False, size=10, cor="000000"):
    return Font(name="Arial", bold=bold, size=size, color=cor)

def _fill(cor):
    return PatternFill(fill_type="solid", fgColor=cor)

def _alinhar(horizontal="left", vertical="center", wrap=False):
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)


# ─── Linha de nutrientes obrigatórios ANVISA ──────────────────────────────────
NUTRIENTES_ROTULO = [
    # (chave_por_100g, chave_por_porcao, chave_vd, nome_exibicao, unidade)
    ("energia_kcal", "energia_kcal", "energia_kcal", "Valor Energético",  "kcal"),
    ("energia_kj",   "energia_kj",   None,           "",                  "kJ"),   # segunda linha do valor energético
    ("carboidrato",  "carboidrato",  "carboidrato",  "Carboidratos",      "g"),
    ("acucares_totais", "acucares_totais", None,      "   Açúcares Totais","g"),
    ("acucares_adicionados","acucares_adicionados","acucares_adicionados","   Açúcares Adicionados","g"),
    ("proteina",     "proteina",     "proteina",     "Proteínas",         "g"),
    ("lipideos",     "lipideos",     "lipideos",     "Gorduras Totais",   "g"),
    ("gordura_saturada","gordura_saturada","gordura_saturada","   Gorduras Saturadas","g"),
    ("gordura_trans","gordura_trans","gordura_trans", "   Gorduras Trans", "g"),
    ("fibra_alimentar","fibra_alimentar","fibra_alimentar","Fibra Alimentar","g"),
    ("sodio",        "sodio",        "sodio",        "Sódio",             "mg"),
]


def gerar_excel(
    nome_produto:      str,
    porcao_gramas:     float,
    num_porcoes:       int | None,
    medida_caseira:    str | None,
    ingredientes_detalhes: list[dict],
    por_100g:          dict,
    por_porcao:        dict,
    vd:                dict,
    peso_total_gramas: float,
) -> bytes:
    """
    Gera arquivo Excel com duas abas e retorna como bytes para download.

    Args:
        nome_produto: nome do produto
        porcao_gramas: tamanho da porção em gramas
        num_porcoes: número de porções por embalagem (opcional)
        medida_caseira: ex: "1 unidade", "2 colheres de sopa" (opcional)
        ingredientes_detalhes: lista com composição individual de cada ingrediente
        por_100g: dict com valores por 100g (já arredondados)
        por_porcao: dict com valores por porção (já arredondados)
        vd: dict com %VD (já calculados)
        peso_total_gramas: peso total da receita em gramas

    Returns:
        bytes do arquivo Excel
    """
    wb = Workbook()

    # ── Aba 1: Tabela Nutricional — Rótulo ─────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Tabela Nutricional — Rótulo"
    _construir_aba_rotulo(ws1, nome_produto, porcao_gramas, num_porcoes,
                           medida_caseira, por_100g, por_porcao, vd)

    # ── Aba 2: Composição Técnica — P&D ─────────────────────────────────────────
    ws2 = wb.create_sheet("Composição Técnica — P&D")
    _construir_aba_pd(ws2, nome_produto, porcao_gramas, ingredientes_detalhes,
                       por_100g, por_porcao, peso_total_gramas)

    # Retornar como bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _construir_aba_rotulo(ws, nome_produto, porcao_gramas, num_porcoes,
                           medida_caseira, por_100g, por_porcao, vd):
    """Constrói a aba da tabela nutricional no padrão ANVISA."""

    # ── Configurar largura das colunas ──────────────────────────────────────────
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 10

    linha = 1

    # ── Linha 1: Nome do Produto ────────────────────────────────────────────────
    ws.merge_cells(f"A{linha}:D{linha}")
    cel = ws.cell(linha, 1, f"Produto: {nome_produto.upper()}")
    cel.font      = _fonte(bold=True, size=11)
    cel.fill      = _fill(COR_HEADER_PD)
    cel.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    cel.alignment = _alinhar("center", "center")
    ws.row_dimensions[linha].height = 20
    linha += 1

    # ── Linha 2: INFORMAÇÃO NUTRICIONAL ────────────────────────────────────────
    ws.merge_cells(f"A{linha}:D{linha}")
    cel = ws.cell(linha, 1, "INFORMAÇÃO NUTRICIONAL")
    cel.font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    cel.fill      = _fill(COR_HEADER_ANVISA)
    cel.alignment = _alinhar("center", "center")
    ws.row_dimensions[linha].height = 24
    linha += 1

    # ── Linha 3: Porção ─────────────────────────────────────────────────────────
    medida = medida_caseira or f"{porcao_gramas:.0f} g"
    texto_porcao = f"Porção de {porcao_gramas:.0f} g ({medida})"
    if num_porcoes:
        texto_porcao += f"  |  {num_porcoes} porção{'ões' if num_porcoes > 1 else ''} por embalagem"

    ws.merge_cells(f"A{linha}:D{linha}")
    cel = ws.cell(linha, 1, texto_porcao)
    cel.font      = _fonte(bold=True, size=10)
    cel.fill      = _fill("EBF3FB")
    cel.alignment = _alinhar("left", "center")
    ws.row_dimensions[linha].height = 18
    linha += 1

    # ── Linha 4: Cabeçalho da tabela ────────────────────────────────────────────
    cabecalhos = ["", f"Por 100 g", f"Por Porção ({porcao_gramas:.0f} g)", "%VD*"]
    for col, txt in enumerate(cabecalhos, 1):
        cel = ws.cell(linha, col, txt)
        cel.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        cel.fill      = _fill(COR_HEADER_ANVISA)
        cel.border    = BORDA_FINA
        cel.alignment = _alinhar("center" if col > 1 else "left", "center")
    ws.row_dimensions[linha].height = 18
    linha += 1

    # ── Linhas de nutrientes ─────────────────────────────────────────────────────
    for i, (ch100, chporcao, chvd, nome_nut, unidade) in enumerate(NUTRIENTES_ROTULO):
        cor_bg = COR_ZEBRA_2 if i % 2 == 0 else COR_ZEBRA_1

        # Coluna A — Nome do nutriente
        bold_row = nome_nut and not nome_nut.startswith("   ")  # sub-itens com indentação
        cel = ws.cell(linha, 1, nome_nut)
        cel.font      = Font(name="Arial", bold=bold_row, size=10)
        cel.fill      = _fill(cor_bg)
        cel.border    = BORDA_FINA
        cel.alignment = _alinhar("left", "center")

        # Coluna B — Por 100g
        val_100g = por_100g.get(ch100, 0)
        if val_100g is None:
            val_100g = 0
        if ch100 == "gordura_trans":
            texto_100g = f"{val_100g:.1f} {unidade}"
        elif unidade == "kcal":
            texto_100g = f"{int(val_100g)} {unidade}"
        elif unidade == "kJ":
            texto_100g = f"{int(val_100g)} {unidade}"
        elif unidade == "mg":
            texto_100g = f"{int(val_100g)} {unidade}"
        else:
            texto_100g = f"{val_100g:.1f} {unidade}"

        cel2 = ws.cell(linha, 2, texto_100g)
        cel2.font      = _fonte(size=10)
        cel2.fill      = _fill(cor_bg)
        cel2.border    = BORDA_FINA
        cel2.alignment = _alinhar("center", "center")

        # Coluna C — Por porção
        val_porcao = por_porcao.get(chporcao, 0)
        if val_porcao is None:
            val_porcao = 0

        # Caso especial: gordura trans
        rotulo_trans = por_porcao.get("gordura_trans_rotulo", None)
        if ch100 == "gordura_trans" and rotulo_trans:
            texto_porcao_cel = rotulo_trans
        elif unidade == "kcal":
            texto_porcao_cel = f"{int(val_porcao)} {unidade}"
        elif unidade == "kJ":
            texto_porcao_cel = f"{int(val_porcao)} {unidade}"
        elif unidade == "mg":
            texto_porcao_cel = f"{int(val_porcao)} {unidade}"
        else:
            texto_porcao_cel = f"{val_porcao:.1f} {unidade}"

        cel3 = ws.cell(linha, 3, texto_porcao_cel)
        cel3.font      = _fonte(size=10)
        cel3.fill      = _fill(cor_bg)
        cel3.border    = BORDA_FINA
        cel3.alignment = _alinhar("center", "center")

        # Coluna D — %VD
        if chvd and chvd in vd and vd[chvd] is not None:
            texto_vd = f"{vd[chvd]}%"
        elif chvd in (None, "gordura_trans", "acucares_totais"):
            texto_vd = "**"
        else:
            texto_vd = "—"

        cel4 = ws.cell(linha, 4, texto_vd)
        cel4.font      = _fonte(size=10)
        cel4.fill      = _fill(cor_bg)
        cel4.border    = BORDA_FINA
        cel4.alignment = _alinhar("center", "center")

        ws.row_dimensions[linha].height = 15
        linha += 1

    # ── Rodapé ──────────────────────────────────────────────────────────────────
    linha += 1
    rodape1 = ("* % Valores Diários de referência com base em uma dieta de 2.000 kcal ou 8.400 kJ. "
               "Seus valores diários podem ser maiores ou menores dependendo de suas necessidades energéticas.")
    ws.merge_cells(f"A{linha}:D{linha}")
    cel = ws.cell(linha, 1, rodape1)
    cel.font      = Font(name="Arial", italic=True, size=8, color="555555")
    cel.alignment = _alinhar("left", "top", wrap=True)
    ws.row_dimensions[linha].height = 30
    linha += 1

    rodape2 = "** Valor Diário não estabelecido."
    ws.merge_cells(f"A{linha}:D{linha}")
    cel = ws.cell(linha, 1, rodape2)
    cel.font      = Font(name="Arial", italic=True, size=8, color="555555")
    cel.alignment = _alinhar("left", "center")
    ws.row_dimensions[linha].height = 14
    linha += 1

    rodape3 = ("⚠️  USO INTERNO — P&D  |  Validar com nutricionista habilitado (CRN) "
               "antes de uso em rótulo comercial  |  NutriCalc P&D")
    ws.merge_cells(f"A{linha}:D{linha}")
    cel = ws.cell(linha, 1, rodape3)
    cel.font      = Font(name="Arial", italic=True, size=8, color="C0392B")
    cel.alignment = _alinhar("center", "center")
    ws.row_dimensions[linha].height = 14


def _construir_aba_pd(ws, nome_produto, porcao_gramas, detalhes, por_100g, por_porcao, peso_total):
    """Constrói a aba de composição técnica detalhada."""
    from modules.database import NUTRIENTES

    NUTRIENTES_LABELS = {
        "umidade": "Umidade (%)", "energia_kcal": "Energia (kcal)",
        "energia_kj": "Energia (kJ)", "proteina": "Proteínas (g)",
        "lipideos": "Gorduras Totais (g)", "colesterol": "Colesterol (mg)",
        "carboidrato": "Carboidratos (g)", "fibra_alimentar": "Fibra Alimentar (g)",
        "cinzas": "Cinzas (g)", "calcio": "Cálcio (mg)", "magnesio": "Magnésio (mg)",
        "manganes": "Manganês (mg)", "fosforo": "Fósforo (mg)", "ferro": "Ferro (mg)",
        "sodio": "Sódio (mg)", "potassio": "Potássio (mg)", "cobre": "Cobre (mg)",
        "zinco": "Zinco (mg)", "retinol": "Retinol (mcg)", "re": "RE (mcg)",
        "rae": "RAE (mcg)", "tiamina": "Tiamina (mg)", "riboflavina": "Riboflavina (mg)",
        "piridoxina": "Piridoxina (mg)", "niacina": "Niacina (mg)",
        "vitamina_c": "Vitamina C (mg)", "vitamina_d": "Vitamina D (mcg)",
        "vitamina_e": "Vitamina E (mg)", "vitamina_b12": "Vitamina B12 (mcg)",
    }

    # Linha de título
    total_cols = 4 + len(NUTRIENTES) * 2  # fixas + nutriente (abs + %)
    ws.merge_cells(f"A1:{get_column_letter(min(total_cols, 30))}1")
    cel = ws.cell(1, 1, f"COMPOSIÇÃO TÉCNICA — {nome_produto.upper()}  |  P&D (USO INTERNO)")
    cel.font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    cel.fill      = _fill(COR_HEADER_PD)
    cel.alignment = _alinhar("center", "center")
    ws.row_dimensions[1].height = 22

    # Colunas fixas: 6 (adicionamos Qtd. Original + Unidade)
    _FIXED = 6
    _NUT_START = _FIXED + 1  # = 7

    # Linha de cabeçalho
    cabecalhos_fixos = [
        "Ingrediente", "Fonte", "Qtd. (g)",
        "Qtd. Original", "Unidade", "% na Receita",
    ]
    cabecalhos_nut = []
    for n in NUTRIENTES:
        label = NUTRIENTES_LABELS.get(n, n)
        cabecalhos_nut.append(label)
        cabecalhos_nut.append("% contr.")

    todos_cab = cabecalhos_fixos + cabecalhos_nut
    for col, txt in enumerate(todos_cab, 1):
        cel = ws.cell(2, col, txt)
        cel.font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        cel.fill      = _fill(COR_HEADER_ANVISA)
        cel.border    = BORDA_FINA
        cel.alignment = _alinhar("center", "center", wrap=True)
    ws.row_dimensions[2].height = 40

    # Configurar larguras
    ws.column_dimensions["A"].width = 30  # Nome
    ws.column_dimensions["B"].width = 12  # Fonte
    ws.column_dimensions["C"].width = 12  # Qtd (g)
    ws.column_dimensions["D"].width = 12  # Qtd Original
    ws.column_dimensions["E"].width = 8   # Unidade
    ws.column_dimensions["F"].width = 13  # % na Receita
    for i in range(len(NUTRIENTES) * 2):
        col_letter = get_column_letter(_NUT_START + i)
        ws.column_dimensions[col_letter].width = 11

    # Calcular totais e percentuais
    peso_total_real = sum(ing["quantidade_gramas"] for ing in detalhes) if detalhes else 1

    # Calcular totais por nutriente
    totais_nut = {n: 0.0 for n in NUTRIENTES}
    for ing in detalhes:
        for n in NUTRIENTES:
            val = ing.get("contribuicao", {}).get(n, 0)
            if val:
                totais_nut[n] += float(val)

    # Linhas de ingredientes
    for i, ing in enumerate(detalhes):
        linha = 3 + i
        cor_bg = COR_ZEBRA_2 if i % 2 == 0 else COR_ZEBRA_1

        pct_receita = (ing["quantidade_gramas"] / peso_total_real * 100) if peso_total_real > 0 else 0
        unidade_orig = ing.get("unidade_original", "g") or "g"
        qtd_orig = ing.get("quantidade_original")
        if qtd_orig is None:
            qtd_orig = ing["quantidade_gramas"]

        dados_fixos = [
            ing.get("nome", "?"),
            ing.get("fonte", "?"),
            round(float(ing["quantidade_gramas"]), 4),
            round(float(qtd_orig), 4),
            unidade_orig,
            f"{pct_receita:.1f}%",
        ]

        for col, val in enumerate(dados_fixos, 1):
            cel = ws.cell(linha, col, val)
            cel.font      = Font(name="Arial", size=9)
            cel.fill      = _fill(cor_bg)
            cel.border    = BORDA_FINA
            cel.alignment = _alinhar("center" if col > 1 else "left", "center")
            if col in (3, 4):  # numeric qty columns: 4-decimal format
                cel.number_format = "0.0000"

        contrib = ing.get("contribuicao", {})
        for j, n in enumerate(NUTRIENTES):
            val_abs  = contrib.get(n, 0) or 0
            total_n  = totais_nut.get(n, 0)
            val_perc = (val_abs / total_n * 100) if total_n > 0 else 0

            col_abs  = _NUT_START + j * 2
            col_perc = _NUT_START + j * 2 + 1

            cel_abs = ws.cell(linha, col_abs, round(float(val_abs), 4))
            cel_abs.font         = Font(name="Arial", size=9)
            cel_abs.fill         = _fill(cor_bg)
            cel_abs.border       = BORDA_FINA
            cel_abs.alignment    = _alinhar("center", "center")
            cel_abs.number_format = "0.0000"

            cel_pct = ws.cell(linha, col_perc, f"{val_perc:.1f}%")
            cel_pct.font      = Font(name="Arial", size=9, color="555555")
            cel_pct.fill      = _fill(cor_bg)
            cel_pct.border    = BORDA_FINA
            cel_pct.alignment = _alinhar("center", "center")

        ws.row_dimensions[linha].height = 14

    # Linha de TOTAIS
    linha_total = 3 + len(detalhes)
    dados_totais = ["TOTAL DA RECEITA", "", round(peso_total_real, 4), "", "", "100%"]
    for col, val in enumerate(dados_totais, 1):
        cel = ws.cell(linha_total, col, val)
        cel.font      = Font(name="Arial", bold=True, size=10)
        cel.fill      = _fill(COR_TOTAL)
        cel.border    = BORDA_FINA
        cel.alignment = _alinhar("center" if col > 1 else "left", "center")
        if col == 3:
            cel.number_format = "0.0000"

    for j, n in enumerate(NUTRIENTES):
        col_abs  = _NUT_START + j * 2
        col_perc = _NUT_START + j * 2 + 1
        val_total = totais_nut.get(n, 0)

        cel_abs = ws.cell(linha_total, col_abs, round(float(val_total), 4))
        cel_abs.font          = Font(name="Arial", bold=True, size=9)
        cel_abs.fill          = _fill(COR_TOTAL)
        cel_abs.border        = BORDA_FINA
        cel_abs.alignment     = _alinhar("center", "center")
        cel_abs.number_format = "0.0000"

        cel_pct = ws.cell(linha_total, col_perc, "100%")
        cel_pct.font   = Font(name="Arial", bold=True, size=9)
        cel_pct.fill   = _fill(COR_TOTAL)
        cel_pct.border = BORDA_FINA
        cel_pct.alignment = _alinhar("center", "center")

    ws.row_dimensions[linha_total].height = 18

    # Linha por 100g
    linha_100g = linha_total + 1
    dados_100g = [f"VALORES POR 100g  (peso total: {round(peso_total_real,1)} g)", "", 100, "", "g", "—"]
    for col, val in enumerate(dados_100g, 1):
        cel = ws.cell(linha_100g, col, val)
        cel.font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        cel.fill      = _fill(COR_HEADER_ANVISA)
        cel.border    = BORDA_FINA
        cel.alignment = _alinhar("center" if col > 1 else "left", "center")

    for j, n in enumerate(NUTRIENTES):
        col_abs = _NUT_START + j * 2
        val_100g = por_100g.get(n, 0) or 0
        cel = ws.cell(linha_100g, col_abs, round(float(val_100g), 4))
        cel.font          = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        cel.fill          = _fill(COR_HEADER_ANVISA)
        cel.border        = BORDA_FINA
        cel.alignment     = _alinhar("center", "center")
        cel.number_format = "0.0000"

        cel2 = ws.cell(linha_100g, col_abs + 1, "—")
        cel2.font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        cel2.fill      = _fill(COR_HEADER_ANVISA)
        cel2.border    = BORDA_FINA
        cel2.alignment = _alinhar("center", "center")

    ws.row_dimensions[linha_100g].height = 16
